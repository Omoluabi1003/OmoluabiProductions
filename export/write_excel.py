from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def _format_sheet(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = header_fill

    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row):
            val = cell[0].value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(80, max(12, max_len + 2))


def write_workbook(
    path: Path,
    master: pd.DataFrame,
    sunbiz_raw: pd.DataFrame,
    irs_raw: pd.DataFrame,
    osm_raw: pd.DataFrame,
    duplicate_review: pd.DataFrame,
    unmatched_review: pd.DataFrame,
    data_dictionary: pd.DataFrame,
) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        master.to_excel(writer, sheet_name="Master", index=False)
        sunbiz_raw.to_excel(writer, sheet_name="Sunbiz Raw", index=False)
        irs_raw.to_excel(writer, sheet_name="IRS Raw", index=False)
        osm_raw.to_excel(writer, sheet_name="OSM Raw", index=False)
        duplicate_review.to_excel(writer, sheet_name="Duplicate Review", index=False)
        unmatched_review.to_excel(writer, sheet_name="Unmatched Review", index=False)
        data_dictionary.to_excel(writer, sheet_name="Data Dictionary", index=False)

        wb = writer.book
        for ws in wb.worksheets:
            _format_sheet(ws)

        master_ws = wb["Master"]
        red = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        orange = PatternFill(start_color="FCE4D6", end_color="F8CBAD", fill_type="solid")
        yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        pink = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")

        headers = {master_ws.cell(row=1, column=i).value: i for i in range(1, master_ws.max_column + 1)}
        for row in range(2, master_ws.max_row + 1):
            sunbiz_status = str(master_ws.cell(row=row, column=headers["Sunbiz Status"]).value or "").lower()
            if sunbiz_status and "active" not in sunbiz_status:
                master_ws.cell(row=row, column=headers["Sunbiz Status"]).fill = red

            irs_rev = str(master_ws.cell(row=row, column=headers["IRS Revocation Status"]).value or "").lower()
            if "revoked" in irs_rev and "not" not in irs_rev:
                master_ws.cell(row=row, column=headers["IRS Revocation Status"]).fill = orange

            score = master_ws.cell(row=row, column=headers["Match Confidence Score"]).value
            try:
                if score is not None and float(score) < 70:
                    master_ws.cell(row=row, column=headers["Match Confidence Score"]).fill = yellow
            except Exception:
                pass

            dup = str(master_ws.cell(row=row, column=headers["Duplicate Flag"]).value or "")
            if dup.strip().lower() == "yes":
                master_ws.cell(row=row, column=headers["Duplicate Flag"]).fill = pink

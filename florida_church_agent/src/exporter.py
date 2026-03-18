"""CSV and Excel export utilities."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from src.deduplicate import DuplicateFlag
from src.models import CleanChurchRecord, RawChurchRecord, SourceResult


def export_raw_csv(records: list[RawChurchRecord], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame([r.model_dump(mode="json") for r in records]).to_csv(path, index=False)


def export_clean_csv(records: list[CleanChurchRecord], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame([r.model_dump(mode="json") for r in records]).to_csv(path, index=False)


def export_logs(
    failed_urls: list[dict],
    source_results: list[SourceResult],
    duplicate_flags: list[DuplicateFlag],
    failed_path: Path,
    source_path: Path,
    dup_path: Path,
) -> None:
    failed_path.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(failed_urls).to_csv(failed_path, index=False)
    pd.DataFrame([s.model_dump(mode="json") for s in source_results]).to_csv(source_path, index=False)
    pd.DataFrame([d.__dict__ for d in duplicate_flags]).to_csv(dup_path, index=False)


def export_excel(
    records: list[CleanChurchRecord],
    source_results: list[SourceResult],
    failed_urls: list[dict],
    out_path: Path,
) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    churches_df = pd.DataFrame([r.model_dump(mode="json") for r in records])
    source_df = pd.DataFrame([s.model_dump(mode="json") for s in source_results])
    failed_df = pd.DataFrame(failed_urls)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        churches_df.to_excel(writer, index=False, sheet_name="Churches")
        source_df.to_excel(writer, index=False, sheet_name="Source Summary")
        failed_df.to_excel(writer, index=False, sheet_name="Failed URLs")

        wb = writer.book
        ws = wb["Churches"]
        ws.freeze_panes = "A2"

        for cell in ws[1]:
            cell.font = Font(bold=True)

        ws.auto_filter.ref = ws.dimensions

        for col_idx, column in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in column:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 60)
